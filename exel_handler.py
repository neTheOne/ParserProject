import os
import pandas as pd
import openpyxl


def create_empty_excel(filename: str):
    '''
    Создание пустого exel фалйа
    :param filename: Нзвание exel файла (filename.xlsx)
    :return: путь к exel файлу
    '''
    os.makedirs("excel_files", exist_ok=True)
    filepath = os.path.join("excel_files", filename)
    wb = openpyxl.Workbook()      # создаёт книгу
    wb.save(filepath)

    return filepath


def add_sheet_and_rows(filepath: str, sheet_name : str, rows: list):
    '''
    Создание нового листа в exel файле и заполнение его из листа
    :param filensme: имя exel-файла
    :param sheet_name: имя листа exel
    :param rows: лист для заполнение листа
    '''
    # открываем файл
    wb = openpyxl.load_workbook(filepath)
    # создаём новый лист
    ws = wb.create_sheet(title=sheet_name)
    for row in rows:
        ws.append(row)

    wb.save(filepath)

def create_excel_from_dict_list(dict_list: list, output_filename: str, sheet_name):
    '''
    функция записывает данные из словаря данных в exel
    :param dict_list: словарь с данными
    :param output_filename: имя файла excel
    :param sheet_name: название страницы
    :return:
    '''
    # Создаем директорию, если она не существует
    if not os.path.exists('excel_files'):
        os.makedirs('excel_files')

    filepath = os.path.join('excel_files', output_filename)

    # Создаем новую книгу Excel
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Записываем данные из списка словарей в Excel
    if dict_list:
        header = list(dict_list[0].keys())
        ws.append(header)  # Записываем заголовки

        for row in dict_list:
            ws.append([row[col] for col in header])

    # Настраиваем стили для красивого вида
    header_style = NamedStyle(name='header')
    header_style.font = Font(bold=True, color='FFFFFF')
    header_style.alignment = Alignment(horizontal='center', vertical='center')
    header_style.fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    border_style = Border(
        left=Side(border_style='thin', color='000000'),
        right=Side(border_style='thin', color='000000'),
        top=Side(border_style='thin', color='000000'),
        bottom=Side(border_style='thin', color='000000')
    )
    header_style.border = border_style

    cell_style = NamedStyle(name='cell')
    cell_style.alignment = Alignment(horizontal='left', vertical='center')
    cell_style.border = border_style

    for cell in ws[1]:  # Применяем стиль к заголовкам
        cell.style = header_style

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.style = cell_style

    # Автоматическое изменение ширины столбцов
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Сохраняем файл
    wb.save(filepath)
    return filepath